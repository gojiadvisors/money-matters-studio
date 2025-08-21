# utils_export.py

import pandas as pd
import json
from datetime import datetime

# --- Expense Grouping ---
EXPENSE_GROUPS = {
    "Essentials": ["Housing", "Utilities", "Food", "Transportation", "Insurance", "Phone/Internet"],
    "Family & Health": ["Childcare", "Health & Wellness", "Education"],
    "Lifestyle": ["Subscriptions", "Discretionary", "Shopping", "Personal Care", "Travel"],
    "Financial Goals": ["Investments", "Giving"],
    "Other": ["Other"]
}

def get_budget_snapshot(session_state):
    # --- FIRE Inputs ---
    annual_income = session_state.get("annual_income", 0)
    annual_savings = session_state.get("annual_savings", 0)
    savings_rate = round(annual_savings / annual_income, 2) if annual_income else 0
    discretionary_spend = annual_income - annual_savings - sum([
        session_state.get(f"{cat}_expense", 0) for cat in session_state.get("expense_categories", [])
    ])

    # --- Lifestyle Selections ---
    household_type = session_state.get("household_type", "N/A")
    location_tier = session_state.get("location_tier", "N/A")
    budget_template = session_state.get("budget_template", "N/A")

    # --- Expense Breakdown ---
    expense_data = []
    for group, categories in EXPENSE_GROUPS.items():
        for category in categories:
            value = session_state.get(f"{category}_expense", None)
            if value is not None:
                expense_data.append({
                    "Group": group,
                    "Category": category,
                    "Monthly Expense ($)": value
                })

    # --- Create DataFrame ---
    df_expenses = pd.DataFrame(expense_data)

    # --- Add Metadata Row ---
    metadata = {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Household Type": household_type,
        "Location Tier": location_tier,
        "Lifestyle Template": budget_template,
        "Annual Income ($)": annual_income,
        "Annual Savings ($)": annual_savings,
        "Savings Rate": savings_rate,
        "Estimated Discretionary Spend ($)": discretionary_spend,
        "User Notes": session_state.get("user_notes", "")
    }

    return df_expenses, metadata

def render_export_buttons(snapshot_tuple):
    import streamlit as st
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from io import BytesIO
    import pandas as pd

    df_expenses, metadata = snapshot_tuple

    # --- Create workbook ---
    wb = Workbook()

    # === Sheet 1: Metadata ===
    ws_meta = wb.active
    ws_meta.title = "Budget Summary"
    ws_meta.append(["Field", "Value"])
    for key, value in metadata.items():
        ws_meta.append([key, value])

    # Style header
    for cell in ws_meta["A1:B1"][0]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    # Auto column widths
    for col in ["A", "B"]:
        ws_meta.column_dimensions[col].width = 30

    # === Sheet 2: Expenses ===
    ws_exp = wb.create_sheet(title="Expense Breakdown")
    ws_exp.append(["Group", "Category", "Monthly Expense ($)"])

    # Write expenses
    for _, row in df_expenses.iterrows():
        ws_exp.append([row["Group"], row["Category"], row["Monthly Expense ($)"]])

    # Add total row
    total = df_expenses["Monthly Expense ($)"].sum()
    ws_exp.append(["TOTAL", "", total])

    # Style header
    for cell in ws_exp["A1:C1"][0]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

    # Format currency column
    for row in ws_exp.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Auto column widths
    for col in range(1, 4):
        col_letter = get_column_letter(col)
        ws_exp.column_dimensions[col_letter].width = 25

    # === Pie Chart: Category Spend ===
    pie = PieChart()
    pie.title = "Spending by Category"

    data_start = 2
    data_end = data_start + len(df_expenses) - 1
    pie_values = Reference(ws_exp, min_col=3, min_row=data_start, max_row=data_end)
    pie_labels = Reference(ws_exp, min_col=2, min_row=data_start, max_row=data_end)
    pie.add_data(pie_values, titles_from_data=False)
    pie.set_categories(pie_labels)
    ws_exp.add_chart(pie, "E2")

    # === Bar Chart: Group Spend ===
    group_totals = df_expenses.groupby("Group")["Monthly Expense ($)"].sum().reset_index()
    ws_chart = wb.create_sheet(title="Charts")
    ws_chart.append(["Group", "Total Monthly Expense ($)"])
    for _, row in group_totals.iterrows():
        ws_chart.append([row["Group"], row["Monthly Expense ($)"]])

    # Style chart sheet
    for cell in ws_chart["A1:B1"][0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for col in ["A", "B"]:
        ws_chart.column_dimensions[col].width = 30

    # Format currency
    for row in ws_chart.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Bar chart
    bar = BarChart()
    bar.title = "Total Spend by Group"
    bar.y_axis.title = "Monthly Expense ($)"
    bar.x_axis.title = "Group"

    bar_data = Reference(ws_chart, min_col=2, min_row=1, max_row=1 + len(group_totals))
    bar_labels = Reference(ws_chart, min_col=1, min_row=2, max_row=1 + len(group_totals))
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_labels)
    bar.shape = 4
    ws_chart.add_chart(bar, "D2")

    # === Save to buffer ===
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # === Download button ===
    st.download_button(
        label="👉 >> 📁 Export to Excel >>",
        data=buffer,
        file_name="lifestyle_budget_snapshot.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )